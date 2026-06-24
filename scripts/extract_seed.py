"""
Comprehensive structured data extraction for Prestige Garage.
Produces JSON files that will seed the Next.js + Prisma database.
"""
import json
import openpyxl
from pathlib import Path
from datetime import datetime, date

UPLOAD_DIR = Path("/home/z/my-project/upload")
OUTPUT = Path("/home/z/my-project/scripts/seed_data.json")


def s(v):
    """Normalize cell value to string/None."""
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float):
        if v.is_integer():
            return int(v)
        return round(v, 4)
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
        return int(f) if f.is_integer() else f
    except (ValueError, TypeError):
        return None


def extract_rolls():
    """Extract all rolls from جرد بروتيكشن file."""
    wb = openpyxl.load_workbook(UPLOAD_DIR / "جرد بروتيكشن يونيو ٢٠٢٦.xlsx", data_only=True)
    rolls = []
    consumptions = []

    # الرولات sheet — header row is row 4
    ws = wb["الرولات"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[4:]:  # start from row 5 (0-indexed 4)
        if not row[0]:
            continue
        code = s(row[0])
        if not code or code.startswith(" ريال") or "كود" in str(code):
            continue
        rolls.append({
            "code": code,
            "brand": s(row[1]),
            "type": s(row[2]),
            "model": s(row[3]),
            "width": num(row[4]),
            "total_length": num(row[5]),
            "price": num(row[6]),
            "supplier": s(row[7]),
            "purchase_date": s(row[8]),
            "notes": s(row[9]),
        })

    # الاستهلاك sheet — header row is row 4
    ws = wb["الاستهلاك"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[4:]:
        if not row[0] and not row[1]:
            continue
        consumptions.append({
            "date": s(row[0]),
            "roll_code": s(row[1]),
            "client_name": s(row[2]),
            "car_type": s(row[3]),
            "plate_number": s(row[4]),
            "meters_used": num(row[5]),
            "usage_area": s(row[6]),
            "waste": num(row[7]),
            "work_order": s(row[8]),
            "notes": s(row[9]),
            "transaction_type": s(row[10]),
        })

    return rolls, consumptions


def extract_employees():
    """Extract employees and June 2026 attendance + advances + commissions."""
    wb = openpyxl.load_workbook(UPLOAD_DIR / "محاسبة موظفين و فنيين  يونيو ٢٠٢٦.xlsx", data_only=True)
    employees = []
    attendance = []
    advances = []
    commissions = []

    # الإعدادات — header row 4, data row 5
    ws = wb["الإعدادات"]
    rows = list(ws.iter_rows(values_only=True))
    emp_cols = None  # will detect from header row
    for i, row in enumerate(rows):
        if row[0] == "اسم الموظف":
            emp_cols = i
            break
    if emp_cols is not None:
        for row in rows[emp_cols + 1:]:
            name = s(row[0])
            if not name or "⛳️" in str(name):
                continue
            employees.append({
                "name": name,
                "base_salary": num(row[1]),
                "phone": s(row[2]),
                "hire_date": s(row[3]),
                "job_title": s(row[4]),
                "notes": s(row[5]),
                "status": "متوقف" if (row[5] and "متوقف" in str(row[5])) else "نشط",
            })

    # June 2026 attendance
    # Row 6 has headers: اليوم | التاريخ | موظف جديد | علي يحيى | أحمد السيد | حسن عبد اللطيف | م/ أمير عمرو
    ws = wb["يونيو 2026"]
    rows = list(ws.iter_rows(values_only=True))
    # Find header row
    header_idx = None
    emp_names = []
    for i, row in enumerate(rows):
        if row[0] == "اليوم" and row[1] == "التاريخ":
            header_idx = i
            # Columns 3..6 are employee names (skip "موظف جديد" col 2)
            for c in row[2:7]:
                if c and "موظف" not in str(c):
                    emp_names.append(s(c))
            break

    if header_idx is not None:
        for row in rows[header_idx + 1:]:
            day = s(row[0])
            date_str = s(row[1])
            if day is None or date_str is None:
                continue
            # Stop when day not numeric
            try:
                int(str(day))
            except (ValueError, TypeError):
                continue
            # Iterate employee cols
            for i, emp_name in enumerate(emp_names):
                status_raw = s(row[3 + i])  # offset: day, date, "موظف جديد", then employees
                if status_raw in ("ح", "غ", "إ", "ر", None):
                    # Build date for June 2026
                    try:
                        day_num = int(str(day))
                        date_iso = f"2026-06-{day_num:02d}"
                    except (ValueError, TypeError):
                        continue
                    attendance.append({
                        "employee_name": emp_name,
                        "date": date_iso,
                        "day_name": date_str,
                        "status": status_raw or "غ",
                        "month": 6,
                        "year": 2026,
                    })

    # Advances (السلفيات) — header row 4, data row 5
    ws = wb["السلفيات"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == "المسلسل":
            header_idx = i
            break
    if header_idx is not None:
        for row in rows[header_idx + 1:]:
            name = s(row[1])
            if not name or "إجمالي" in str(name):
                continue
            advances.append({
                "employee_name": name,
                "date": s(row[2]),
                "amount": num(row[3]),
                "notes": s(row[4]),
                "month": num(row[5]),
                "year": num(row[6]),
            })

    # Services log from employees file — extract commissions
    # Row 5 has the actual header
    ws = wb["سجل الخدمات"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(rows):
        if row[0] == "#" and row[1] == "التاريخ":
            header_idx = i
            break
    if header_idx is not None:
        for row in rows[header_idx + 1:]:
            tech_name = s(row[3])  # اسم الفني
            if not tech_name:
                continue
            commissions.append({
                "date": s(row[1]),
                "month": s(row[2]),
                "technician": tech_name,
                "client_name": s(row[4]),
                "car_type": s(row[5]),
                "service_type": s(row[6]),
                "amount": num(row[7]),
                "notes": s(row[8]),
                "service_category": s(row[9]),
            })

    return employees, attendance, advances, commissions


def extract_stock_services_invoices():
    """Extract stock, services, invoices from prestige_garage_v3 file."""
    wb = openpyxl.load_workbook(UPLOAD_DIR / "prestige_garage_v3 (1).xlsx", data_only=True)
    stock_items = []
    stock_movements = []
    services = []
    invoices = []

    # مخزون_الدتيلنج — header row 2
    ws = wb["مخزون_الدتيلنج"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:
        first = s(row[0])
        if first is None or "اجمالي" in str(first) or "📌" in str(first):
            continue
        name = s(row[1])
        if not name:
            continue
        stock_items.append({
            "name": name,
            "category": "ديتيلنج",
            "unit": s(row[2]),
            "total_received": num(row[3]),
            "total_withdrawn": num(row[4]),
            "current_qty": num(row[5]),
            "min_level": num(row[6]),
            "status": s(row[7]),
            "unit_price": num(row[8]),
        })

    # مخزون_البوليش
    ws = wb["مخزون_البوليش"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:
        first = s(row[0])
        if first is None or "اجمالي" in str(first) or "📌" in str(first):
            continue
        name = s(row[1])
        if not name:
            continue
        stock_items.append({
            "name": name,
            "category": "بوليش وكوتينج",
            "unit": s(row[2]),
            "total_received": num(row[3]),
            "total_withdrawn": num(row[4]),
            "current_qty": num(row[5]),
            "min_level": num(row[6]),
            "status": s(row[7]),
            "unit_price": num(row[8]),
        })

    # استلام_الخامات — header row 2
    ws = wb["استلام_الخامات"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[2:]:
        first = s(row[0])
        if first is None or "اجمالي" in str(first) or "📌" in str(first):
            continue
        name = s(row[2])
        if not name:
            continue
        stock_movements.append({
            "date": s(row[1]),
            "item_name": name,
            "material_type": s(row[3]),
            "movement_type": s(row[4]),
            "quantity": num(row[5]),
            "unit": s(row[6]),
            "unit_price": num(row[7]),
            "total_cost": num(row[8]),
            "notes": s(row[9]),
            "delivery_note": s(row[10]),
        })

    # سجل_الخدمات — header row 3
    ws = wb["سجل_الخدمات"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[3:]:
        op_code = s(row[0])
        if not op_code or "اجمالي" in str(op_code):
            continue
        services.append({
            "code": op_code,
            "date": s(row[1]),
            "plate": s(row[2]),
            "client_name": s(row[3]),
            "car_type": s(row[4]),
            "service_type": s(row[5]),
            "price": num(row[6]),
            "payment_method": s(row[7]),
            "technician": s(row[8]),
            "notes": s(row[9]),
        })

    # ملخص_الفواتير — header row 4
    ws = wb["ملخص_الفواتير"]
    rows = list(ws.iter_rows(values_only=True))
    for row in rows[4:]:
        dn = s(row[0])
        if not dn or "اجمالي" in str(dn) or "***" in str(dn):
            continue
        invoices.append({
            "delivery_note": dn,
            "date": s(row[1]),
            "description": s(row[2]),
            "total": num(row[3]),
            "discount": num(row[4]),
            "net": num(row[5]),
            "items_count": num(row[6]),
            "notes": s(row[7]),
        })

    return stock_items, stock_movements, services, invoices


def main():
    print("Extracting rolls inventory...")
    rolls, consumptions = extract_rolls()
    print(f"  Rolls: {len(rolls)}, Consumptions: {len(consumptions)}")

    print("Extracting employees data...")
    employees, attendance, advances, commissions = extract_employees()
    print(f"  Employees: {len(employees)}, Attendance: {len(attendance)}, Advances: {len(advances)}, Commissions: {len(commissions)}")

    print("Extracting stock & services & invoices...")
    stock_items, stock_movements, services, invoices = extract_stock_services_invoices()
    print(f"  Stock items: {len(stock_items)}, Movements: {len(stock_movements)}, Services: {len(services)}, Invoices: {len(invoices)}")

    data = {
        "rolls": rolls,
        "consumptions": consumptions,
        "employees": employees,
        "attendance": attendance,
        "advances": advances,
        "commissions": commissions,
        "stock_items": stock_items,
        "stock_movements": stock_movements,
        "services": services,
        "invoices": invoices,
    }

    OUTPUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSaved to: {OUTPUT}")
    print(f"Total records: {sum(len(v) for v in data.values())}")

    # Sample preview
    print("\n=== SAMPLE PREVIEW ===")
    print("First roll:", json.dumps(rolls[0] if rolls else None, ensure_ascii=False))
    print("First employee:", json.dumps(employees[0] if employees else None, ensure_ascii=False))
    print("First service:", json.dumps(services[0] if services else None, ensure_ascii=False))
    print("First stock item:", json.dumps(stock_items[0] if stock_items else None, ensure_ascii=False))
    print("First invoice:", json.dumps(invoices[0] if invoices else None, ensure_ascii=False))


if __name__ == "__main__":
    main()
