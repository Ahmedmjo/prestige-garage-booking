"""
Extract actual data from the 3 Prestige Garage Excel files.
Output: /home/z/my-project/scripts/extracted_data.json
"""
import json
import openpyxl
from pathlib import Path
from datetime import datetime, date

UPLOAD_DIR = Path("/home/z/my-project/upload")
OUTPUT_FILE = Path("/home/z/my-project/scripts/extracted_data.json")


def cell_to_str(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip() if isinstance(v, str) else v


def read_sheet_data(wb, sheet_name, header_row=1, max_rows=None):
    """Read sheet starting from header_row, return list of dicts."""
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Find header row
    if header_row - 1 >= len(rows):
        return [], []
    headers_raw = rows[header_row - 1]
    headers = [cell_to_str(h) for h in headers_raw]

    # Start from header_row + 1
    data_rows = rows[header_row:]
    if max_rows:
        data_rows = data_rows[:max_rows]

    result = []
    for row in data_rows:
        # Skip empty rows
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                key = headers[i] if headers[i] else f"col_{i}"
                row_dict[key] = cell_to_str(val)
        if any(v is not None and v != "" for v in row_dict.values()):
            result.append(row_dict)
    return headers, result


def extract_rolls_inventory():
    """جرد بروتيكشن يونيو ٢٠٢٦.xlsx"""
    file_path = UPLOAD_DIR / "جرد بروتيكشن يونيو ٢٠٢٦.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)

    data = {"rolls": [], "consumptions": [], "balance": [], "financial": []}

    # الرولات sheet - rolls master
    # First find the actual header row
    ws = wb["الرولات"]
    rows = list(ws.iter_rows(values_only=True))
    # print first 5 rows for inspection
    print("=== الرولات first 5 rows ===")
    for i, r in enumerate(rows[:5]):
        print(f"Row {i+1}: {r[:10]}")

    # الاستهلاك sheet - consumptions
    ws2 = wb["الاستهلاك"]
    rows2 = list(ws2.iter_rows(values_only=True))
    print("\n=== الاستهلاك first 5 rows ===")
    for i, r in enumerate(rows2[:5]):
        print(f"Row {i+1}: {r}")

    return wb, file_path


def extract_employees():
    """محاسبة موظفين و فنيين  يونيو ٢٠٢٦.xlsx"""
    file_path = UPLOAD_DIR / "محاسبة موظفين و فنيين  يونيو ٢٠٢٦.xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)

    print("\n\n=== EMPLOYEES FILE ===")
    # الإعدادات - employees master
    ws = wb["الإعدادات"]
    rows = list(ws.iter_rows(values_only=True))
    print("\n--- الإعدادات sheet (all rows) ---")
    for i, r in enumerate(rows):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r}")

    # يونيو 2026 - June payroll
    ws2 = wb["يونيو 2026"]
    rows2 = list(ws2.iter_rows(values_only=True))
    print("\n--- يونيو 2026 sheet (first 30 rows) ---")
    for i, r in enumerate(rows2[:30]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r}")

    # سجل الخدمات - service log (first 10 rows for structure)
    ws3 = wb["سجل الخدمات"]
    rows3 = list(ws3.iter_rows(values_only=True))
    print("\n--- سجل الخدمات sheet (first 5 rows, first 20 cols) ---")
    for i, r in enumerate(rows3[:5]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:20]}")

    # السلفيات - advances
    ws4 = wb["السلفيات"]
    rows4 = list(ws4.iter_rows(values_only=True))
    print("\n--- السلفيات sheet (first 10 rows) ---")
    for i, r in enumerate(rows4[:10]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r}")

    return wb, file_path


def extract_prestige_v3():
    """prestige_garage_v3 (1).xlsx"""
    file_path = UPLOAD_DIR / "prestige_garage_v3 (1).xlsx"
    wb = openpyxl.load_workbook(file_path, data_only=True)

    print("\n\n=== PRESTIGE V3 FILE ===")

    # لوحة_التحكم
    ws = wb["لوحة_التحكم"]
    rows = list(ws.iter_rows(values_only=True))
    print("\n--- لوحة_التحكم (first 30 rows) ---")
    for i, r in enumerate(rows[:30]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:13]}")

    # مخزون_الدتيلنج
    ws2 = wb["مخزون_الدتيلنج"]
    rows2 = list(ws2.iter_rows(values_only=True))
    print("\n--- مخزون_الدتيلنج (first 30 rows) ---")
    for i, r in enumerate(rows2[:30]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:10]}")

    # مخزون_البوليش
    ws3 = wb["مخزون_البوليش"]
    rows3 = list(ws3.iter_rows(values_only=True))
    print("\n--- مخزون_البوليش (first 30 rows) ---")
    for i, r in enumerate(rows3[:30]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:10]}")

    # الأدوات_والمعدات
    ws4 = wb["الأدوات_والمعدات"]
    rows4 = list(ws4.iter_rows(values_only=True))
    print("\n--- الأدوات_والمعدات (first 30 rows) ---")
    for i, r in enumerate(rows4[:30]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:11]}")

    # استلام_الخامات
    ws5 = wb["استلام_الخامات"]
    rows5 = list(ws5.iter_rows(values_only=True))
    print("\n--- استلام_الخامات (first 10 rows) ---")
    for i, r in enumerate(rows5[:10]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:11]}")

    # ملخص_الفواتير
    ws6 = wb["ملخص_الفواتير"]
    rows6 = list(ws6.iter_rows(values_only=True))
    print("\n--- ملخص_الفواتير (all rows) ---")
    for i, r in enumerate(rows6):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:8]}")

    # سجل_الخدمات
    ws7 = wb["سجل_الخدمات"]
    rows7 = list(ws7.iter_rows(values_only=True))
    print("\n--- سجل_الخدمات (first 15 rows) ---")
    for i, r in enumerate(rows7[:15]):
        if any(c is not None for c in r):
            print(f"Row {i+1}: {r[:15]}")

    return wb, file_path


if __name__ == "__main__":
    print("=" * 80)
    print("EXTRACTING DATA FROM PRESTIGE GARAGE EXCEL FILES")
    print("=" * 80)

    wb1, _ = extract_rolls_inventory()
    wb2, _ = extract_employees()
    wb3, _ = extract_prestige_v3()

    print("\n\n=== EXTRACTION COMPLETE ===")
